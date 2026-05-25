import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL  = PatternFill("solid", fgColor="1a6fa8")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
ODD_FILL     = PatternFill("solid", fgColor="F4F8FC")
THIN         = Side(style="thin", color="D0DDE8")
CELL_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def export_to_excel(data: dict, path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Title
    ws["A1"] = data.get("title", "Report")
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)
    ws["A2"] = data.get("subtitle", "")
    ws["A2"].font = Font(name="Calibri", italic=True, color="6B8299", size=10)
    ws.append([])

    columns = data.get("columns", [])
    rows    = data.get("rows", [])

    # Header row
    ws.append(columns)
    header_row_idx = ws.max_row
    for cell in ws[header_row_idx]:
        cell.fill        = HEADER_FILL
        cell.font        = HEADER_FONT
        cell.alignment   = Alignment(horizontal="center", vertical="center")
        cell.border      = CELL_BORDER

    # Data rows
    for idx, row in enumerate(rows):
        ws.append(row)
        fill = ODD_FILL if idx % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
        for cell in ws[ws.max_row]:
            cell.fill      = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = CELL_BORDER

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

    ws.row_dimensions[header_row_idx].height = 22
    ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)

    wb.save(path)
