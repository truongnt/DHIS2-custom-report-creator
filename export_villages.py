"""
Export village org units (name contains "B.") with parent and grandparent info.
Output: villages_export.xlsx
"""
import sys
import getpass
import requests
from requests.auth import HTTPBasicAuth
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# Load saved profiles from Windows Credential Manager
try:
    sys.path.insert(0, str(__import__("pathlib").Path(__file__).parent))
    from config.credentials import load_profiles, load_password, profile_label
    _CRED_OK = True
except Exception:
    _CRED_OK = False


def fetch_villages(base_url: str, username: str, password: str):
    auth = HTTPBasicAuth(username, password)
    session = requests.Session()
    session.auth = auth

    url = base_url.rstrip("/") + "/api/organisationUnits.json"
    params = {
        "filter": ["name:like:B.", "path:!like:KEGktGHjhYQ"],
        "fields": "id,name,displayName,geometry,parent[id,name,displayName,parent[id,name,displayName]]",
        "paging": "false",
    }
    print("Fetching org units...")
    resp = session.get(url, params=params, timeout=120)
    resp.raise_for_status()
    data = resp.json()
    return data.get("organisationUnits", [])


def export_excel(rows, output_path: str):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Villages"

    headers = [
        "Village UID", "Village Name", "Village Display Name",
        "Longitude", "Latitude",
        "Parent UID", "Parent Name", "Parent Display Name",
        "Grandparent UID", "Grandparent Name", "Grandparent Display Name",
    ]

    header_fill = PatternFill("solid", fgColor="1a6fa8")
    header_font = Font(bold=True, color="FFFFFF")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r, ou in enumerate(rows, 2):
        parent = ou.get("parent") or {}
        grandparent = parent.get("parent") or {}
        geom = ou.get("geometry") or {}
        coords = geom.get("coordinates") if geom.get("type") == "Point" else None
        lon = coords[0] if coords else ""
        lat = coords[1] if coords else ""
        ws.append([
            ou.get("id", ""),
            ou.get("name", ""),
            ou.get("displayName", ""),
            lon,
            lat,
            parent.get("id", ""),
            parent.get("name", ""),
            parent.get("displayName", ""),
            grandparent.get("id", ""),
            grandparent.get("name", ""),
            grandparent.get("displayName", ""),
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    wb.save(output_path)
    print(f"Saved {len(rows)} rows → {output_path}")


def main():
    base_url = username = password = ""

    if _CRED_OK:
        profiles = load_profiles()
        if profiles:
            print("Saved profiles:")
            for i, p in enumerate(profiles):
                print(f"  [{i+1}] {profile_label(p)}")
            choice = input(f"Choose profile [1-{len(profiles)}] or Enter to type manually: ").strip()
            if choice.isdigit() and 1 <= int(choice) <= len(profiles):
                p = profiles[int(choice) - 1]
                base_url = p["url"]
                username = p["username"]
                password = load_password(base_url, username)
                print(f"Using: {username} @ {base_url}")

    if not base_url:
        base_url = input("DHIS2 URL [https://hmis.gov.la/hmis]: ").strip() or "https://hmis.gov.la/hmis"
    if not username:
        username = input("Username: ").strip()
    if not password:
        password = getpass.getpass("Password: ")

    try:
        ous = fetch_villages(base_url, username, password)
    except requests.HTTPError as e:
        print(f"HTTP error: {e}")
        sys.exit(1)

    ous.sort(key=lambda x: x.get("name", ""))
    print(f"Found {len(ous)} villages")

    output = r"d:\Projects CHAI\1. Laos\Auto report\villages_export.xlsx"
    export_excel(ous, output)


if __name__ == "__main__":
    main()
